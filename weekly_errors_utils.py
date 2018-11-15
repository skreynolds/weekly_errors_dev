#! /usr/bin/python


# Import the modules that are required for the script
import openpyxl #module used to interact with excel spreadsheets
from sheet_field_dict import sheet_field_dict, sheet_field_dict_margaret


######################################################################################
# Function to import excel workbooks into Python
######################################################################################
def import_excel(file):
	return openpyxl.load_workbook(file)


######################################################################################
# Function checks the integrity of the imported worksheets
######################################################################################
def file_integrity_check(wb, error_flag):
	
	print('Checking file integrity...')

	# Integrity indicator
	integrity = True

	# Get a list of the sheet names and check to ensure all
	# sheets are present
	sheet_names = wb.sheetnames

	if error_flag == 'douglas':
		# Get a list of the 
		key_names = sheet_field_dict.keys()
		try:
			assert sorted(list(sheet_names)) == sorted(list(key_names))
		except:
			print('The sheet names of the file you are trying to import do not match the required structure. Check the documentation.')
			print('[Not O.K.]\n')
			integrity = False
			return integrity
	else:
		# Get a list of the 
		key_names = sheet_field_dict_margaret.keys()
		try:
			assert sorted(list(sheet_names)) == sorted(list(key_names))
		except:
			print('The sheet names of the file you are trying to import do not match the required structure. Check the documentation.')
			print('[Not O.K.]\n')
			integrity = False
			return integrity


	for sheet in sheet_names:

		# Store the current sheet
		sheet_object = wb[sheet]

		# Extract the mandated fields from the source dictionary
		if error_flag == 'douglas':
			sheet_fields = sheet_field_dict[sheet]
		else:
			sheet_fields = sheet_field_dict_margaret[sheet]

		for i,field in enumerate(sheet_fields, start=1):
			
			try:
				assert sheet_object.cell(row=1, column=i).value == field
			except:
				print('Sheet Name:', sheet)
				print(	'Source Truth:', field,
						'Import:', sheet_object.cell(row=1, column=i).value, '\n')
				integrity = False
	
	if integrity == True:
		print('[O.K.]\n')
	else:
		print('[Not O.K.]\n')

	return integrity


######################################################################################
# Function checks criteria is the same from the markup and new error spreadsheet
######################################################################################
def row_match_check(new_sheet, markup_sheet, current_row, current_markup_row, col_list):
	for col_i in col_list:
		if (not new_sheet.cell(row=current_row, column=col_i).value 
				and not markup_sheet.cell(row=current_markup_row, column=col_i).value):
			continue
		elif (new_sheet.cell(row=current_row, column=col_i).value 
				!= markup_sheet.cell(row=current_markup_row, column=col_i).value):
			return False
	return True


######################################################################################
# Function fills cells with colour for given column list
######################################################################################
def fill_cells(sheet, colour, current_row, col_list):
	
	# Iterate through the column and colour each individually
	for col_i in col_list:
		sheet.cell(row=current_row, column=col_i).fill = \
						openpyxl.styles.PatternFill(fill_type="solid", start_color=colour, end_color=colour)


######################################################################################
# Function checks that the routine has not left rows unaccounted 
######################################################################################
def check_routine(sheet, process_count):
	if process_count == sheet.max_row:
		print('[O.K.]\n')
	else:
		print('[Not O.K.]\n')


######################################################################################
# Function checks that there are no blank comments in the markup file
######################################################################################
def markup_complete_check(markup_wb):

	comments_complete = True
	sheets = markup_wb.sheetnames

	for sheet_i in sheets:
		markup_sheet = markup_wb[sheet_i]
		max_row_markup = markup_sheet.max_row

		for row_i in range(2,max_row_markup+1):
			if not (markup_sheet.cell(row=row_i, column=1).value):
				comments_complete = False

	if comments_complete == True:
		print('[O.K.]\n')
	else:
		print('[Not O.K.]\n')

	return comments_complete


######################################################################################
# Performs the routine on the enrols_2018_TLS_TL0 (douglas)
######################################################################################
def enrols_2018_TLS_TL0(markup_sheet, new_sheet):

	process_count = 1
	max_row_new = new_sheet.max_row
	max_col_new = new_sheet.max_column
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [4, 6, 8]):
				
				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break
		
		if not new_sheet.cell(row=row_new, column=1).value:
			process_count += 1
			fill_cells(new_sheet, "FFFF0000", row_new, list(range(2,max_col_new+1)))

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the act_end_dt_error (douglas)
######################################################################################
def act_end_dt_error(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 5, 6, 8, 10]):
				
				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			process_count += 1
			fill_cells(new_sheet, "FFFF0000", row_new, [10])

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the act_st_dt_error (douglas)
######################################################################################
def act_st_dt_error(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 4, 5, 7]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			process_count += 1
			fill_cells(new_sheet, "FFFF0000", row_new, [8])

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the dup_students (douglas)
######################################################################################
def dup_students(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			continue

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the student_dups_2017_18 (douglas)
######################################################################################
def student_dups_2017_18(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_col_new = new_sheet.max_column
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 7]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			process_count += 1
			fill_cells(new_sheet, "FFFF0000", row_new, list(range(2,max_col_new+1)))
			
	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the vfh_errors (douglas)
######################################################################################
def vfh_errors(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [2, 4, 5, 6, 7]):
				
				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			
			# Store the error type for convenience
			error_type = new_sheet.cell(row=row_new, column=2).value
			
			###############################################################################################################
			# Highlight fields for Both Home Location Postcode and Home Location Country Code in HE Stats cannot be null
			###############################################################################################################
			if error_type == "Both Home Location Postcode and Home Location Country Code in HE Stats cannot be null":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"FFFF0000",row_new, [2,22,23])
			
			###############################################################################################################
			# Highlight fields for Both Term Location Postcode and Term Location Country Code in HE Stats cannot be null
			###############################################################################################################
			elif error_type == "Both Term Location Postcode and Term Location Country Code in HE Stats cannot be null":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00CCFFCC",row_new, [2,20,21])

			###############################################################################################################
			# Highlight fields for Course, Unit and/or Teaching period does not correspond to the VFH Unit schedule
			###############################################################################################################
			elif error_type == "Course, Unit and/or Teaching period does not correspond to the VFH Unit schedule":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00FFFF00",row_new, [2,4,6,7])
			
			###############################################################################################################
			# Highlight fields for VFH enrolment but missing CHESSN
			###############################################################################################################
			elif error_type == "VFH enrolment but missing CHESSN":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00800080",row_new, [2,14])
			
			###############################################################################################################
			# Highlight fields for VFH enrolment but missing OR invalid ATSI Code
			###############################################################################################################
			elif error_type == "VFH enrolment but missing OR invalid ATSI Code":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"0000FF00",row_new, [2,16])

			###############################################################################################################
			# Highlight fields for VFH enrolment but missing OR invalid Birth Country Code
			###############################################################################################################
			elif error_type == "VFH enrolment but missing OR invalid Birth Country Code":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00CCFFFF",row_new, [2,18])

			###############################################################################################################
			# Highlight fields for VFH enrolment but missing OR invalid Citizenship Status
			###############################################################################################################
			elif error_type == "VFH enrolment but missing OR invalid Citizenship Status":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00993366",row_new, [2,17])

			###############################################################################################################
			# Highlight fields for VFH enrolment but missing OR invalid Highest Attainment Code
			###############################################################################################################
			elif error_type == "VFH enrolment but missing OR invalid Highest Attainment Code":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00FF6600",row_new, [2,28])

			###############################################################################################################
			# Highlight fields for VFH enrolment but missing OR invalid Highest Attainment Yr
			###############################################################################################################
			elif error_type == "VFH enrolment but missing OR invalid Highest Attainment Yr":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00333300",row_new, [2,29])

			###############################################################################################################
			# Highlight fields for VFH enrolment but missing OR invalid Unit Student Status
			###############################################################################################################
			elif error_type == "VFH enrolment but missing OR invalid Unit Student Status":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00FFFFCC",row_new, [2,32])

			###############################################################################################################
			# Highlight fields for VFH enrolment but missing OR invalid VFH Language Code
			###############################################################################################################
			elif error_type == "VFH enrolment but missing OR invalid VFH Language Code":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00C0C0C0",row_new, [2,19])

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the enr_2017_comp_2018 (douglas)
######################################################################################
def enr_2017_comp_2018(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [2, 6, 8]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			continue

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the general_errors (douglas)
######################################################################################
def general_errors(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			'''
			##################################################################
			if (new_sheet.cell(row=row_new, column=2).value == "Student is missing a verified USI. Please correct."
			and markup_sheet.cell(row=row_markup, column=2).value == "Student is missing a verified USI. Please correct."):
				print("Row " + str(row_new) + " in NEW and Row " + str(row_markup) + " in MARKUP")
				print(new_sheet.cell(row=row_new, column=25).value,
						new_sheet.cell(row=row_new, column=26).value,
						markup_sheet.cell(row=row_markup, column=25).value,
						markup_sheet.cell(row=row_markup, column=26).value)
				print(new_sheet.cell(row=row_new, column=25).value == markup_sheet.cell(row=row_markup, column=25).value)
				print(not new_sheet.cell(row=row_new, column=25).value)
			#################################################################
			'''
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [2, 6, 25, 26]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			
			# Store the error type for convenience
			error_type = new_sheet.cell(row=row_new, column=2).value

			###############################################################################################################
			# Highlight fields for Advanced Standing missing Basis Details. Please check.
			###############################################################################################################
			if error_type == "Advanced Standing missing Basis Details. Please check.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00CCFFCC",row_new, [2])

			###############################################################################################################
			# Highlight fields for ATSI student born overseas. Please check
			###############################################################################################################
			elif error_type == "ATSI student born overseas. Please check":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00FFFF00",row_new, [2,9,15])

			###############################################################################################################
			# Highlight fields for ATSI student with language other than English or Indigenous. Please check.
			###############################################################################################################
			elif error_type == "ATSI student with language other than English or Indigenous. Please check.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00FF0000",row_new, [2,9,14])

			###############################################################################################################
			# Highlight fields for Diploma and above course enrolment but not in VFH Teaching period. Please check.
			###############################################################################################################
			elif error_type == "Diploma and above course enrolment but not in VFH Teaching period. Please check.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00CC99FF",row_new, [2,5])

			###############################################################################################################
			# Highlight fields for ETP Course commencing this year but now LAPSED or DISCONTIN. Should this still be under ETP?
			###############################################################################################################
			elif error_type == "ETP Course commencing this year but now LAPSED or DISCONTIN. Should this still be under ETP?":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00FF00FF",row_new, [2,27])

			###############################################################################################################
			# Highlight fields for Doesn't meet ETP eligibility criteria - too young, at school, or non-NT resident. Please check.
			###############################################################################################################
			elif error_type == "Doesn't meet ETP eligibility criteria - too young, at school, or non-NT resident. Please check.":
				pass

			###############################################################################################################
			# Highlight fields for Funding Source is Redundant or Absent
			###############################################################################################################
			elif error_type == "Funding Source is Redundant or Absent":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00003366", row_new, [2,27])

			###############################################################################################################
			# Highlight fields for Invalid Residential Postcode. Cannot have a post box postcode.
			###############################################################################################################
			elif error_type == "Invalid Residential Postcode. Cannot have a post box postcode":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"009999FF", row_new, [2,16])

			###############################################################################################################
			# Highlight fields for Invalid State.
			###############################################################################################################
			elif error_type == "Invalid State":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00333333", row_new, [2,17])

			###############################################################################################################
			# Highlight fields for At School Flag N but VETIS Flag Y
			###############################################################################################################
			elif error_type == "At School Flag N but VETIS Flag Y":
				process_count += 1
				fill_cells(new_sheet, "00FF00FF", row_new, [2,20,21])

			###############################################################################################################
			# Highlight fields for LRNSUPP Student with At School Flag Y. Please confirm if they are still at school
			###############################################################################################################
			elif error_type == "LRNSUPP Student with At School Flag Y. Please confirm if they are still at school":
				process_count += 1
				fill_cells(new_sheet, "00993366", row_new, [2,20,25])

			###############################################################################################################
			# Highlight fields for Missing or invalid Study Reason ID
			###############################################################################################################
			elif error_type == "Missing or invalid Study Reason ID":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00008000",row_new, [2,31])

			###############################################################################################################
			# Highlight fields for Student has 30A funding code but a home postcode different to OSPC. Please check.
			###############################################################################################################
			elif error_type == "Student has 30A funding code but a home postcode different to OSPC. Please check.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00CCFFFF",row_new, [2,16,27])
						
			###############################################################################################################
			# Highlight fields for Student Home Language is Australian Indigenous but student birth country not Australia. Please check.
			###############################################################################################################
			elif error_type == "Student Home Language is Australian Indigenous but student birth country not Australia. Please check.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00CCFFFF",row_new, [2,14,15])

			###############################################################################################################
			# Highlight fields for Student is missing a verified USI. Please correct.
			###############################################################################################################
			elif error_type == "Student is missing a verified USI. Please correct.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"0033CCCC",row_new, [2,40])

			###############################################################################################################
			# Highlight fields for VETIS funding code but VETIS Flag N
			###############################################################################################################
			elif error_type == "VETIS funding code but VETIS Flag N":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00008080",row_new, [2,21,27])

			###############################################################################################################
			# Highlight fields for Student has At School Flag Y but is older than usual. Please check.
			###############################################################################################################
			elif error_type == "Student has At School Flag Y but is older than usual. Please check.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00008080",row_new, [2,11,20])

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the course_intent_errors (douglas)
######################################################################################
def course_intent_errors(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [2, 6, 25, 26]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			
			# Store the error type for convenience
			error_type = new_sheet.cell(row=row_new, column=2).value

			###############################################################################################################
			# Highlight fields for Student has a course intention conflict. Cannot say yes to more than one. Please check.
			###############################################################################################################
			if error_type == "Student has a course intention conflict. Cannot say yes to more than one. Please check.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"00CCFFCC",row_new, [2, 32, 33, 34])

			###############################################################################################################
			# Highlight fields for Course intention details are missing. Please check.
			###############################################################################################################
			elif error_type == "Course intention details are missing. Please check.":
				# Highlight errors
				process_count += 1
				fill_cells(new_sheet,"000000FF",row_new, [2, 32, 33, 34])

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the inconsis_fee_fund_scr_pls_chk (douglas)
######################################################################################
def inconsis_fee_fund_scr_pls_chk(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [2, 6, 8]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			continue

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the duplicate_sua (douglas)
######################################################################################
def duplicate_sua(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 4, 6, 8, 12]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			continue

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the superseded_unit (douglas)
######################################################################################
def superseded_unit(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 4, 5, 6, 7, 8, 11]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			continue

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on the vfh_w_chk_for_participation (douglas)
######################################################################################
def vfh_w_chk_for_participation(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_col_new = new_sheet.max_column
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 6, 8]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			process_count += 1
			fill_cells(new_sheet, "FFFF0000", row_new, list(range(2,max_col_new+1)))

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on TCI-not J (margaret)
######################################################################################
def tci_not_j(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_col_new = new_sheet.max_column
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 4, 5, 6, 7]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			process_count += 1
			fill_cells(new_sheet, "FFFF0000", row_new, list(range(2,max_col_new+1)))

	check_routine(new_sheet, process_count)


######################################################################################
# Performs the routine on K Missing TCI (margaret)
######################################################################################
def k_missing_tci(markup_sheet, new_sheet):
	
	process_count = 1
	max_row_new = new_sheet.max_row
	max_col_new = new_sheet.max_column
	max_row_markup = markup_sheet.max_row

	for row_new in range(2, max_row_new+1):
		for row_markup in range(2, max_row_markup+1):
			if row_match_check(new_sheet, markup_sheet, row_new, row_markup, [3, 4, 5, 6]):

				process_count += 1
				new_sheet.cell(row=row_new, column=1).value = markup_sheet.cell(row=row_markup, column=1).value
				break

		if not new_sheet.cell(row=row_new, column=1).value:
			process_count += 1
			fill_cells(new_sheet, "FFFF0000", row_new, list(range(2,max_col_new+1)))

	check_routine(new_sheet, process_count)